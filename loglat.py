"""
Program to plot the coordinates of a location on google map.
"""
import pandas
import requests
import gmplot
from openpyxl import load_workbook

# Reading the file contaning list of cities and their states
# Downlaod from www.downloadexcelfiles.com/wo_en/download-list-cities-canada
df = pandas.read_csv('list-cities-canada.csv')
# Combing Cities and states in form of tuples
address = zip(list(df['City']), list(df['State']))
# Google API Key
api_key = "***************************************"
# Variable for storing latitude and longitude
longlat = []
# A counter to check the number of locaions passed to google api
count = 0
# Working with Google Geocode Api
for state in address:
    api_response = requests.get('https://maps.googleapis.com/maps/api/geocode/json?address={0}&key={1}'.format(state, api_key))
    api_response_dict = api_response.json()
    if api_response_dict['status'] == 'OK':
        latitude = api_response_dict['results'][0]['geometry']['location']['lat']
        longitude = api_response_dict['results'][0]['geometry']['location']['lng']
        count += 1
        print(count)
        print('State: {0} Latitude: {1} Longitude: {2}'.format(state, latitude, longitude))
        longlat.append((state, latitude, longitude))

# Saving all the longitude and longitude in an Excel File
df = pandas.DataFrame(longlat, columns=["City/State", "Latitude", "Longitude"])
book = load_workbook('coordinates.xlsx')
writer = pandas.ExcelWriter('coordinates.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#  Essentially these steps are just loading the existing data from 'Masterfile.xlsx' and populating your writer with them.
df.to_excel(writer, "USA")
writer.save()

# Working with TJX Api
# Reading the Excel file containing Coordinates
df2 = pandas.read_excel('coordinates.xlsx', sheet_name='Canada')
latitude = list(df2['Latitude'])
longitude = list(df2['Longitude'])
longlat = zip(latitude, longitude)
store_address = []
print("Finding address....")
count = 0
for coordinates in longlat:
    count += 1
    print(count)
    tjx_api_url = "http://mktsvc.tjx.com/storelocator/GetSearchResults?geolat={0}&geolong={1}&chain=chain=21%2C20&radius=100".format(coordinates[0], coordinates[1])
    tjx_api_api_response = requests.get(tjx_api_url)
    tjx_api_response_dict = tjx_api_api_response.json()
    print(tjx_api_response_dict)
    for store in tjx_api_response_dict["Stores"]:
        addr = store.get("Address", 'NA')
        addr2 = store.get("Address2", "NA")
        city = store.get("City", "NA")
        st = store.get("State", "NA")
        zipcode = store.get("Zip", "NA")
        country = store.get("Country", "NA")
        storeid = store.get("StoreID", "NA")
        hours = store.get("Hours", "NA")
        lat = store.get("Latitude", "NA")
        longco = store.get("Longitude", "NA")
        name = store.get("Name", "NA")
        phone = store.get("Phone", "NA")
        store_address.append((addr, addr2, city, st, zipcode, country, storeid, hours, lat, longco, name, phone))
        print(store_address)

df = pandas.DataFrame(store_address, columns=["Address", "Address2", "City", "State", "Zip", "Country", "StoreID", "Hours", "Latitude", "Longitude", "Name", "Phone"])
book = load_workbook('stores.xlsx')
writer = pandas.ExcelWriter('stores.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#  Essentially these steps are just loading the existing data from 'Masterfile.xlsx' and populating your writer with them.
df.to_excel(writer, "Canada")
writer.save()

df_ca = pandas.read_excel('stores.xlsx', sheet_name="Canada")
# Removing duplicate coordinates
l = list((set((list(zip(list(df_ca["Latitude"]), list(df_ca["Longitude"])))))))
ll = list(zip(*l))
latitude = ll[0]
longitude = ll[1]
# Mapping Coordinates on Google map.
# Initialize the map to the first location in the list
gmap = gmplot.GoogleMapPlotter(latitude[0], longitude[0], 4)
gmap.heatmap(latitude, longitude, radius=15,)
# Write the map in an HTML file
gmap.draw('europe.html')
